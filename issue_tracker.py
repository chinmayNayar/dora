"""
Parse Book2.xlsx issue tracker for DORA Metric 3 (CFR) and Metric 4 (MTTR).

Rule (per ops sheet):
  Issue Observed = Yes  →  rollback issued  →  counts as change failure.
"""

from __future__ import annotations

import logging
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)

DEFAULT_SHEET = Path(__file__).resolve().parent / "Book2.xlsx"

# Primary row columns (sheet has no header row)
COL_SNO = 0
COL_CHG = 1
COL_PROJECT = 2
COL_ISSUE = 7
COL_ROLLBACK = 8
COL_START = 9
COL_END = 10
COL_STATUS = 11
COL_DESC = 15
COL_INC = 17

_MONTHS = {
    "jan": 1, "january": 1, "feb": 2, "february": 2, "mar": 3, "march": 3,
    "apr": 4, "april": 4, "may": 5, "jun": 6, "june": 6, "jul": 7, "july": 7,
    "aug": 8, "august": 8, "sep": 9, "sept": 9, "september": 9,
    "oct": 10, "october": 10, "nov": 11, "november": 11, "dec": 12, "december": 12,
}


def _yes(val: Any) -> bool:
    s = str(val or "").strip().lower()
    return s in ("yes", "yes.", "y", "true", "1")


def _parse_date(val: Any) -> Optional[datetime]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    s = str(val).strip()
    if not s:
        return None
    # 12 May, 2026 | 22nd June, 2026 | 02 June, 2026
    s = re.sub(r"(\d+)(st|nd|rd|th)\b", r"\1", s, flags=re.I)
    s = s.replace(",", " ")
    m = re.match(r"^(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})$", s.strip())
    if m:
        day, mon, year = int(m.group(1)), m.group(2).lower(), int(m.group(3))
        mi = _MONTHS.get(mon)
        if mi:
            try:
                return datetime(year, mi, day)
            except ValueError:
                return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s[:10], fmt)
        except ValueError:
            continue
    return None


def _extract_chgs(val: Any) -> List[str]:
    if val is None:
        return []
    return re.findall(r"CHG\d+", str(val).upper())


def _extract_incs(val: Any) -> List[str]:
    if val is None:
        return []
    return re.findall(r"INC\d+", str(val).upper())


# Same Est. Time (minutes) as Pipeline tab Stage-by-Stage Breakdown (PD[].em)
PIPELINE_STAGE_EM_MIN = [
    8,   # Pipeline Execution
    14,  # 3Scale Configuration
    10,  # Azure Key Vault
    5,   # ArgoCD Sync
    10,  # GitHub Changes
    10,  # Branch Conflict
    15,  # Jenkins Setup
    120, # New Service Setup
    10,  # Multi-pipeline Race
]


def pipeline_mttr_hours() -> float:
    """MTTR from pipeline Est. Time only (not CMR open→closed dates)."""
    total_min = sum(PIPELINE_STAGE_EM_MIN)
    return round(total_min / 60.0, 2) if total_min > 0 else 0.0


def _mttr_hours(start: Optional[datetime], end: Optional[datetime]) -> float:
    # Deprecated for Metric 4 — kept unused; MTTR comes from pipeline Est. Time.
    return pipeline_mttr_hours()


def load_issue_rows(path: Optional[Path] = None) -> List[Dict[str, Any]]:
    """Load primary CMR rows from the issue tracker workbook."""
    path = Path(path) if path else DEFAULT_SHEET
    if not path.exists():
        logger.warning("Issue sheet not found: %s", path)
        return []

    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.error("openpyxl not installed — cannot read %s", path)
        return []

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows: List[Dict[str, Any]] = []

    for r in range(1, (ws.max_row or 0) + 1):
        vals = [ws.cell(r, c).value for c in range(1, 19)]
        chgs = _extract_chgs(vals[COL_CHG])
        if not chgs:
            continue
        # Skip continuation-only lines without a project/sno-like primary marker
        if vals[COL_PROJECT] is None and vals[COL_SNO] is None:
            continue

        issue = _yes(vals[COL_ISSUE])
        rollback = _yes(vals[COL_ROLLBACK])
        start = _parse_date(vals[COL_START])
        end = _parse_date(vals[COL_END])
        # Failure rule: Issue Observed = Yes (rollback accompanies it)
        is_failure = issue
        desc = str(vals[COL_DESC] or "").strip()
        incs = _extract_incs(vals[COL_INC]) + _extract_incs(vals[COL_DESC])
        project = str(vals[COL_PROJECT] or "").strip() or "Unknown"
        date_str = (start or end).strftime("%Y-%m-%d") if (start or end) else ""

        rows.append({
            "sno": vals[COL_SNO],
            "chgs": chgs,
            "project": project,
            "issue_observed": issue,
            "rollback": rollback or issue,  # issue yes implies rollback per ops rule
            "is_failure": is_failure,
            "start": start.isoformat() if start else "",
            "end": end.isoformat() if end else "",
            "date": date_str,
            "status": str(vals[COL_STATUS] or "").strip(),
            "description": desc,
            "incidents": sorted(set(incs)),
            "mttr_hours": pipeline_mttr_hours() if is_failure else 0.0,
            "sheet_row": r,
        })

    logger.info("Loaded %s issue-tracker rows from %s (%s failures)",
                len(rows), path.name, sum(1 for x in rows if x["is_failure"]))
    return rows


def build_chg_overlay(rows: Optional[List[Dict[str, Any]]] = None) -> Dict[str, Dict[str, Any]]:
    """Map CHG number → failure/MTTR overlay."""
    rows = rows if rows is not None else load_issue_rows()
    out: Dict[str, Dict[str, Any]] = {}
    for row in rows:
        for chg in row["chgs"]:
            prev = out.get(chg)
            # Prefer failure=True if any row marks it
            if prev and prev.get("is_failure") and not row["is_failure"]:
                continue
            out[chg] = {
                "is_failure": row["is_failure"],
                "rollback": row["rollback"],
                "mttr_hours": row["mttr_hours"],
                "project": row["project"],
                "description": row["description"],
                "incidents": row["incidents"],
                "date": row["date"],
                "status": row["status"],
            }
    return out


def apply_issue_overlay(dashboard: Dict[str, Any],
                        sheet_path: Optional[Path] = None) -> Dict[str, Any]:
    """
    Merge Book2.xlsx failures with live ServiceNow close_code failures.
    Does not erase API close_code marks — sheet failures are additive.
    """
    rows = load_issue_rows(sheet_path)
    overlay = build_chg_overlay(rows)
    extras = dashboard.get("cmr_extra") or []
    cmr = dashboard.get("cmr_data") or []
    snow_incs = list(dashboard.get("incidents") or [])
    restore_h = pipeline_mttr_hours()

    matched = 0
    if overlay:
        for i, extra in enumerate(extras):
            chg = str(extra.get("chg") or "").upper()
            if not chg or chg not in overlay:
                continue
            info = overlay[chg]
            if i >= len(cmr):
                continue
            # Additive: keep close_code failures; OR in sheet Issue Observed=Yes
            if info["is_failure"]:
                cmr[i]["i"] = True
                cmr[i]["r"] = bool(cmr[i].get("r")) or bool(info["rollback"])
                cmr[i]["m"] = float(info["mttr_hours"] or restore_h or 0)
                matched += 1
            elif info.get("rollback"):
                cmr[i]["r"] = True

    # Sheet failure incidents (additive with ServiceNow close_code incidents)
    sheet_incs: List[Dict[str, Any]] = []
    seen = {str(x.get("ch") or "").upper() for x in snow_incs}
    for row in rows:
        if not row["is_failure"]:
            continue
        key = tuple(row["chgs"])
        # skip if any of these CHGs already listed from ServiceNow
        if any(c in seen for c in row["chgs"]):
            continue
        for c in row["chgs"]:
            seen.add(c)
        try:
            dt_disp = datetime.strptime(row["date"], "%Y-%m-%d").strftime("%d %b %Y") if row["date"] else ""
        except ValueError:
            dt_disp = row["date"]
        inc_nums = ", ".join(row["incidents"]) if row["incidents"] else ""
        issue_txt = row["description"] or f"Issue observed on {row['chgs'][0]} — rollback issued"
        if inc_nums:
            issue_txt = f"{issue_txt} [{inc_nums}]".strip()
        sheet_incs.append({
            "p": row["project"],
            "t": "inc",
            "ch": " / ".join(row["chgs"]),
            "dt": dt_disp,
            "is": issue_txt,
            "sv": "e",
            "rb": True,
            "mttr_hours": row["mttr_hours"] or restore_h,
            "source": "book2_issue_observed",
        })

    dashboard["cmr_data"] = cmr
    dashboard["incidents"] = snow_incs + sheet_incs
    dashboard["issue_sheet"] = {
        "loaded": bool(overlay),
        "path": str(sheet_path or DEFAULT_SHEET.name),
        "rows": len(rows),
        "failures": sum(1 for r in rows if r["is_failure"]),
        "matched_cmrs": matched,
        "sheet_only_incidents": len(sheet_incs),
        "rule": (
            "Metric 3 = ServiceNow close_code (Unsuccessful | Successful with issues) "
            "OR Book2 Issue Observed=Yes; MTTR = Pipeline Stage Est. Time sum"
        ),
        "mttr_source": "pipeline_stage_est_time",
        "mttr_hours": restore_h,
        "mttr_minutes": sum(PIPELINE_STAGE_EM_MIN),
        "close_code_stats": dashboard.get("close_code_stats"),
    }
    return dashboard
